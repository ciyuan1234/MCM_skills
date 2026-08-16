#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
附件要求解析器 (make-requirements.py) —— 从赛题 PDF 自动提取"需提交结果文件"要求
用法:
    python make-requirements.py <工作目录> [赛题PDF/文本路径]
示例:
    python make-requirements.py D:\\2026_国赛1234
    python make-requirements.py . 0_赛题\\C题.pdf

输出:
    - 5_支撑材料/requirements.json  要求清单（结果文件名/模板来源/对应问题/说明）
    - 5_支撑材料/填写引导.md         填写引导（模板结构、行列表头、填数要点）
    - 5_支撑材料/结果模板/           官方模板文件（在 0_赛题/本地资料库/1_数据 中定位并复制）；
                                      找不到官方模板时按题面信息生成占位模板 xlsx

解析线索（官方赛题常见措辞）:
    - "将结果填入 result1_1.xlsx" / "结果分别填入 result1_1.xlsx 和 result1_2.xlsx"
    - "模板文件见附件3" / "（模板文件见附件 3）"
    - "附件1 乡村现有耕地和农作物的基本情况"
"""

import glob
import json
import os
import re
import sys


def read_text(path):
    with open(path, "rb") as f:
        raw = f.read()
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def extract_problem_text(problem_file):
    """从赛题 PDF/文本提取全文。"""
    if not problem_file or not os.path.isfile(problem_file):
        return ""
    low = problem_file.lower()
    if low.endswith(".pdf"):
        try:
            from pypdf import PdfReader
            parts = []
            for p in PdfReader(problem_file).pages:
                try:
                    parts.append(p.extract_text() or "")
                except Exception:
                    pass
            return "\n".join(parts)
        except Exception as e:
            print(f"[警告] PDF 解析失败: {e}，尝试按文本读取")
    return read_text(problem_file)


def find_problem_file(workdir):
    """在 0_赛题 下找赛题文件（pdf/txt/docx）。"""
    d = os.path.join(workdir, "0_赛题")
    if not os.path.isdir(d):
        return None
    for f in sorted(os.listdir(d)):
        low = f.lower()
        if low.endswith((".pdf", ".txt", ".docx")) and "format" not in low:
            return os.path.join(d, f)
    return None


def parse_requirements(text, workdir, problem_file):
    """解析要求清单。"""
    if not text.strip():
        return [], "赛题文本为空，无法解析"
    reqs = []
    # 1) "将结果填入 X.xlsx" / "结果分别填入 A.xlsx 和 B.xlsx" / "数值结果填入附件A"
    for m in re.finditer(r"(?:将|把)(?:[^。；\n]{0,60}?)?结果(?:分别)?填入([^。；\n]{0,60})", text):
        seg = m.group(1)
        files = re.findall(r"(?:[\w\u4e00-\u9fa5\-]+\.xlsx?|附件\s*[A-Z\u4e00-\u9fa5]{1,6})", seg)
        for f in files:
            f = f.replace(" ", "")
            if not any(r["文件"] == f for r in reqs):
                reqs.append({"文件": f, "模板": None, "问题": [], "说明": seg.strip()[:80]})
    # 2) 单独文件名要求（result*.xlsx 模式，未带"填入"动词的补充）
    for m in re.finditer(r"result[\w\-]*\.xlsx|附件[\u4e00-\u9fa5A-Z]?\s*[\u4e00-\u9fa5]{0,10}(?:数据结果|结果|模板)\.xlsx", text):
        f = m.group(0).strip()
        if not any(r["文件"] == f for r in reqs):
            reqs.append({"文件": f, "模板": None, "问题": [], "说明": ""})
    # 3) 模板引用 "模板文件见附件X" / "（模板见附件 3）"
    for m in re.finditer(r"模板(?:文件)?\s*(?:见|在)\s*附件\s*([一二三四五六七八九十\d]+)", text):
        n = m.group(1)
        for r in reqs:
            if not r["模板"]:
                r["模板"] = f"附件{n}"
    # 4) 问题归属: 找"问题X"标题与其后 500 字符内的结果文件名
    cur_q = None
    for m in re.finditer(r"问题\s*([一二三四五六七八九十1-9]+)", text):
        q = m.group(1)
        window = text[m.end():m.end() + 500]
        files = re.findall(r"(?:[\w\u4e00-\u9fa5\-]+\.xlsx?|附件\s*[A-Z\u4e00-\u9fa5]{1,6})", window)
        for f in files:
            f = f.replace(" ", "")
            for r in reqs:
                if r["文件"] == f and q not in r["问题"]:
                    r["问题"].append(q)
    return reqs, ""


def locate_templates(workdir, reqs, problem_file):
    """在 0_赛题 / 1_数据 / 本地资料库中定位官方模板文件。"""
    bases = []
    for d in ("0_赛题", "1_数据"):
        p = os.path.join(workdir, d)
        if os.path.isdir(p):
            bases.append(p)
    src = os.path.join(os.path.dirname(problem_file or ""), ".")
    if src and os.path.isdir(src):
        bases.append(src)
    # 本地资料库（若存在）
    for lib in (r"D:\全国大学生数学竞赛资料",):
        if os.path.isdir(lib):
            bases.append(lib)
    for r in reqs:
        name = r["文件"]
        pattern = name + "*.xlsx" if not name.lower().endswith(".xlsx") else name
        for b in bases:
            hits = glob.glob(os.path.join(b, "**", pattern), recursive=True)
            if hits:
                r["模板路径"] = hits[0]
                break
        else:
            r.setdefault("模板路径", None)
    return reqs


def make_placeholder(workdir, r, text):
    """找不到官方模板时生成占位模板（含题面上下文表头线索）。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = r["文件"].replace(".xlsx", "").replace(" ", "_")[:31]
    headers = []
    # 从题面找该文件附近的表头线索（如"供应商ID, 周"）——宽松：截取该文件名后 200 字
    m = re.search(re.escape(r["文件"]), text)
    if m:
        seg = text[m.end():m.end() + 200]
        cand = re.findall(r"[\u4e00-\u9fa5A-Za-z]{2,}(?:ID|编号|周|月|年|供应商|转运商|序号|名称)", seg)
        headers = list(dict.fromkeys(cand))[:6]
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    out = os.path.join(workdir, "5_支撑材料", "结果模板", r["文件"])
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    r["模板路径"] = out
    r["占位模板"] = True


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    workdir = sys.argv[1]
    problem_file = sys.argv[2] if len(sys.argv) > 2 else find_problem_file(workdir)
    text = extract_problem_text(problem_file)
    reqs, err = parse_requirements(text, workdir, problem_file)
    if err:
        print(f"[错误] {err}")
        sys.exit(1)
    if not reqs:
        print("[警告] 未解析到结果文件要求（赛题可能无附件结果文件，如纯理论题）")
        reqs = []
    reqs = locate_templates(workdir, reqs, problem_file)
    for r in reqs:
        if not r.get("模板路径"):
            make_placeholder(workdir, r, text)
    # 写 requirements.json
    sup = os.path.join(workdir, "5_支撑材料")
    os.makedirs(sup, exist_ok=True)
    payload = {"赛题": os.path.basename(problem_file or ""), "结果文件": reqs}
    with open(os.path.join(sup, "requirements.json"), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    # 写填写引导
    lines = ["# 附件填写引导", "",
             f"- 赛题文件: {os.path.basename(problem_file or '(未找到)')}",
             f"- 解析出结果文件要求: {len(reqs)} 个", ""]
    for r in reqs:
        lines += [f"## {r['文件']}",
                  f"- 对应问题: {('、'.join(r['问题']) if r['问题'] else '未解析到明确问题归属')}",
                  f"- 模板: {r.get('模板路径') or '(未找到官方模板，已生成占位模板)'}",
                  f"- 说明: {r['说明'] or '(无)'}",
                  ""]
    lines += ["## 填写要求（skill 红线）", "",
              "1. 结果必须由 2_代码 的脚本按 模板结构 填入（openpyxl 保持模板格式，如 fill_attachments.py 模式）",
              "2. 先运行脚本读取模板（load_workbook 保留原 sheet 结构与格式），只填数值单元格",
              "3. 数值来源必须是 results_*.csv 的实际输出，禁止手写/编造",
              "4. 填写后的成品放入 5_支撑材料（与 requirements.json 同名对应），并在论文附录文件列表中列出",
              "5. 完成后运行 checks.py + verify.py 确认无回归", ""]
    with open(os.path.join(sup, "填写引导.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"[完成] requirements.json / 填写引导.md / 结果模板 已写入 {sup}")
    for r in reqs:
        print(f"  - {r['文件']}  <-- {r.get('模板路径')}")


if __name__ == "__main__":
    main()